"""
Appeal Configuration API routes (Admin Service)
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, List
from sqlalchemy.orm import Session
from sqlalchemy import desc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import uuid
import logging

from database.connection import get_admin_db
from database.models import Application, AppealRecord
from services.appeal_service import appeal_service
from utils.i18n_loader import get_translation

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/config", tags=["appeal-config"])


def get_current_user_and_application_from_request(request: Request, db: Session) -> tuple:
    """Get current user and application ID from request context"""
    # Check for X-Application-ID header (highest priority - from frontend selector)
    header_app_id = request.headers.get('x-application-id') or request.headers.get('X-Application-ID')
    if header_app_id:
        try:
            header_app_uuid = uuid.UUID(str(header_app_id))
            app = db.query(Application).filter(
                Application.id == header_app_uuid,
                Application.is_active == True
            ).first()
            if app:
                auth_context = getattr(request.state, 'auth_context', None)
                if auth_context:
                    return auth_context['data'], str(app.id)
        except (ValueError, AttributeError):
            pass

    auth_context = getattr(request.state, 'auth_context', None)
    if not auth_context:
        raise HTTPException(status_code=401, detail="Not authenticated")

    current_user = auth_context['data']

    # Try to get application_id from auth context (new API keys)
    application_id = current_user.get('application_id')
    if application_id:
        return current_user, str(application_id)

    # Fallback: get tenant_id and find default application
    tenant_id = current_user.get('tenant_id')
    if not tenant_id:
        raise HTTPException(status_code=401, detail="Tenant ID not found in auth context")

    try:
        tenant_uuid = uuid.UUID(str(tenant_id))
        default_app = db.query(Application).filter(
            Application.tenant_id == tenant_uuid,
            Application.is_active == True
        ).first()

        if not default_app:
            raise HTTPException(status_code=404, detail="No active application found for user")

        return current_user, str(default_app.id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid tenant ID format")


def get_default_message_template(language: str = 'en') -> str:
    """Get default message template based on language"""
    return get_translation(language, 'appealPage', 'defaultMessageTemplate')


class AppealConfigUpdate(BaseModel):
    """Appeal configuration update model"""
    enabled: bool = Field(False, description="Whether to enable appeal feature")
    message_template: Optional[str] = Field(
        None,
        description="Template for appeal message, {appeal_url} will be replaced with actual URL"
    )
    appeal_base_url: str = Field(
        "",
        description="Base URL for appeal links (e.g., https://domain.com or http://192.168.1.100:5001)"
    )
    final_reviewer_email: Optional[str] = Field(
        None,
        description="Email address for human final review when AI rejects appeal"
    )


class AppealConfigResponse(BaseModel):
    """Appeal configuration response model"""
    id: Optional[str] = None
    enabled: bool
    message_template: str
    appeal_base_url: str
    final_reviewer_email: Optional[str] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


class AppealRecordResponse(BaseModel):
    """Appeal record response model"""
    id: str
    request_id: str
    user_id: Optional[str]
    original_content: str
    original_risk_level: str
    original_categories: List[str]
    status: str
    ai_approved: Optional[bool]
    ai_review_result: Optional[str]
    processor_type: Optional[str] = None
    processor_id: Optional[str] = None
    processor_reason: Optional[str] = None
    created_at: Optional[str]
    ai_reviewed_at: Optional[str]
    processed_at: Optional[str] = None


class ManualReviewRequest(BaseModel):
    """Manual review request model"""
    action: str = Field(..., description="Review action: 'approve' or 'reject'")
    reason: Optional[str] = Field(None, description="Optional reason for the decision")


class AppealRecordsListResponse(BaseModel):
    """Paginated appeal records response"""
    items: List[AppealRecordResponse]
    total: int
    page: int
    page_size: int
    pages: int


@router.get("/appeal", response_model=AppealConfigResponse)
async def get_appeal_config(request: Request, db: Session = Depends(get_admin_db)):
    """Get appeal configuration for current application"""
    try:
        current_user, application_id = get_current_user_and_application_from_request(request, db)

        # Get language preference from Accept-Language header
        accept_language = request.headers.get('accept-language', 'en')
        language = 'zh' if 'zh' in accept_language.lower() else 'en'

        config = await appeal_service.get_config(application_id, db)

        if not config:
            # Return default config if not exists (with i18n message template)
            return AppealConfigResponse(
                enabled=False,
                message_template=get_default_message_template(language),
                appeal_base_url="",
                final_reviewer_email=None
            )

        return AppealConfigResponse(**config)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get appeal config: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get appeal config: {str(e)}")


@router.put("/appeal", response_model=AppealConfigResponse)
async def update_appeal_config(
    config_data: AppealConfigUpdate,
    request: Request,
    db: Session = Depends(get_admin_db)
):
    """Update appeal configuration for current application"""
    try:
        current_user, application_id = get_current_user_and_application_from_request(request, db)
        tenant_id = current_user.get('tenant_id')

        if not tenant_id:
            raise HTTPException(status_code=400, detail="Tenant ID not found")

        config = await appeal_service.update_config(
            application_id=application_id,
            tenant_id=str(tenant_id),
            config_data=config_data.dict(),
            db=db
        )

        return AppealConfigResponse(**config)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update appeal config: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update appeal config: {str(e)}")


@router.get("/appeal/records", response_model=AppealRecordsListResponse)
async def get_appeal_records(
    request: Request,
    status: Optional[str] = Query(None, description="Filter by status: pending, reviewing, pending_review, approved, rejected"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Page size"),
    db: Session = Depends(get_admin_db)
):
    """Get appeal records for current application"""
    try:
        current_user, application_id = get_current_user_and_application_from_request(request, db)

        result = await appeal_service.get_appeal_records(
            application_id=application_id,
            status=status,
            page=page,
            page_size=page_size,
            db=db
        )

        return AppealRecordsListResponse(**result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get appeal records: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get appeal records: {str(e)}")


@router.post("/appeal/records/{appeal_id}/review")
async def manual_review_appeal(
    appeal_id: str,
    review_data: ManualReviewRequest,
    request: Request,
    db: Session = Depends(get_admin_db)
):
    """
    Manually review an appeal (approve or reject)

    This endpoint allows human reviewers to:
    - Approve appeals that were pending human review
    - Override previous decisions (e.g., reject a previously approved appeal)
    """
    try:
        current_user, application_id = get_current_user_and_application_from_request(request, db)

        # Get reviewer email from current user
        reviewer_email = current_user.get('email', 'unknown@unknown.com')

        # Validate action
        if review_data.action not in ['approve', 'reject']:
            raise HTTPException(
                status_code=400,
                detail="Invalid action. Must be 'approve' or 'reject'"
            )

        # Get language from Accept-Language header or default to 'zh'
        accept_language = request.headers.get('Accept-Language', 'zh')
        language = 'zh' if 'zh' in accept_language else 'en'

        result = await appeal_service.manual_review_appeal(
            appeal_id=appeal_id,
            action=review_data.action,
            reviewer_email=reviewer_email,
            reason=review_data.reason,
            language=language,
            db=db
        )

        if not result.get('success'):
            raise HTTPException(
                status_code=400,
                detail=result.get('message', 'Manual review failed')
            )

        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to process manual review: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process manual review: {str(e)}")


@router.get("/appeal/records/export")
async def export_appeal_records(
    request: Request,
    status: Optional[str] = Query(None, description="Filter by status: pending, reviewing, pending_review, approved, rejected"),
    db: Session = Depends(get_admin_db)
):
    """Export appeal records to Excel"""
    try:
        current_user, application_id = get_current_user_and_application_from_request(request, db)

        # Build query
        app_uuid = uuid.UUID(application_id)
        query = db.query(AppealRecord).filter(
            AppealRecord.application_id == app_uuid
        )

        if status:
            query = query.filter(AppealRecord.status == status)

        # Get all results (limit to 10000 for safety)
        records = query.order_by(desc(AppealRecord.created_at)).limit(10000).all()

        # Get application names
        app_ids = set(record.application_id for record in records if record.application_id)
        app_names = {}
        if app_ids:
            apps = db.query(Application).filter(Application.id.in_(app_ids)).all()
            app_names = {app.id: app.name for app in apps}

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Appeal Records"

        # Define headers
        headers = [
            "Request ID",
            "Application",
            "Appeal User",
            "Original Content",
            "Original Risk Level",
            "Original Categories",
            "Status",
            "AI Approved",
            "AI Review Result",
            "Processor Type",
            "Processor ID",
            "Processor Reason",
            "Appeal Time",
            "AI Review Time",
            "Process Time"
        ]

        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Status translation map
        status_map = {
            'pending': 'Pending',
            'reviewing': 'Reviewing',
            'pending_review': 'Pending Review',
            'approved': 'Approved',
            'rejected': 'Rejected'
        }

        # Write data
        for row_num, record in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=record.request_id)
            ws.cell(row=row_num, column=2, value=app_names.get(record.application_id, '-'))
            ws.cell(row=row_num, column=3, value=record.user_id or '-')
            ws.cell(row=row_num, column=4, value=record.original_content[:500] if record.original_content else '')
            ws.cell(row=row_num, column=5, value=record.original_risk_level or '-')
            ws.cell(row=row_num, column=6, value=", ".join(record.original_categories or []))
            ws.cell(row=row_num, column=7, value=status_map.get(record.status, record.status))
            ws.cell(row=row_num, column=8, value="Yes" if record.ai_approved else ("No" if record.ai_approved is False else "-"))
            ws.cell(row=row_num, column=9, value=record.ai_review_result or '-')
            ws.cell(row=row_num, column=10, value=record.processor_type or '-')
            ws.cell(row=row_num, column=11, value=record.processor_id or '-')
            ws.cell(row=row_num, column=12, value=record.processor_reason or '-')
            ws.cell(row=row_num, column=13, value=record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else '-')
            ws.cell(row=row_num, column=14, value=record.ai_reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if record.ai_reviewed_at else '-')
            ws.cell(row=row_num, column=15, value=record.processed_at.strftime('%Y-%m-%d %H:%M:%S') if record.processed_at else '-')

        # Adjust column widths
        column_widths = [30, 20, 20, 60, 15, 30, 15, 12, 50, 15, 20, 40, 20, 20, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename with timestamp
        filename = f"appeal_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Return as streaming response
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export appeal records error: {e}")
        raise HTTPException(status_code=500, detail="Failed to export appeal records")
