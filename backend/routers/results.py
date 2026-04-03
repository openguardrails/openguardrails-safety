from typing import List, Optional, Tuple
import json
import uuid
from pathlib import Path
from io import BytesIO
from datetime import datetime, timedelta, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, text, cast
from sqlalchemy.dialects.postgresql import JSONB
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from database.connection import get_admin_db
from database.models import DetectionResult, Tenant, Application, Workspace
from models.responses import DetectionResultResponse, PaginatedResponse
from utils.logger import setup_logger
from utils.url_signature import generate_signed_media_url
from config import settings

logger = setup_logger()
router = APIRouter(tags=["Results"])

def get_current_tenant(request: Request, db: Session) -> Tenant:
    """Get current tenant from request context"""
    auth_context = getattr(request.state, 'auth_context', None)
    if not auth_context or 'data' not in auth_context:
        raise HTTPException(status_code=401, detail="Not authenticated")

    data = auth_context['data']
    tenant_id = data.get('tenant_id')
    if not tenant_id:
        raise HTTPException(status_code=401, detail="Tenant ID not found in auth context")

    try:
        tenant_uuid = uuid.UUID(str(tenant_id))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid tenant ID format")

    tenant = db.query(Tenant).filter(Tenant.id == tenant_uuid).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    return tenant


def _build_app_workspace_cache(db: Session, tenant_id) -> dict:
    """Build a lookup cache for application names and workspace info"""
    apps = db.query(Application).filter(Application.tenant_id == str(tenant_id)).all()
    workspaces = db.query(Workspace).filter(Workspace.tenant_id == str(tenant_id)).all()

    ws_map = {str(ws.id): ws for ws in workspaces}
    app_map = {}
    for app in apps:
        ws_id = str(app.workspace_id) if getattr(app, 'workspace_id', None) else None
        ws = ws_map.get(ws_id) if ws_id else None
        app_map[str(app.id)] = {
            "application_name": app.name,
            "workspace_id": ws_id,
            "workspace_name": ws.name if ws else None,
        }
    return app_map


def _enrich_result(result, app_map: dict, image_urls: list, truncate_content: bool = True) -> DetectionResultResponse:
    """Convert a DetectionResult to DetectionResultResponse with app/workspace info"""
    app_id_str = str(result.application_id) if result.application_id else None
    app_info = app_map.get(app_id_str, {}) if app_id_str else {}

    content = result.content
    if truncate_content and len(content) > 200:
        content = content[:200] + "..."

    original_content = getattr(result, 'original_content', None)
    has_data_masking = original_content is not None

    return DetectionResultResponse(
        id=result.id,
        request_id=result.request_id,
        content=content,
        original_content=original_content if not truncate_content else None,
        has_data_masking=has_data_masking,
        suggest_action=result.suggest_action,
        suggest_answer=result.suggest_answer,
        hit_keywords=result.hit_keywords,
        model_response=result.model_response if hasattr(result, 'model_response') else None,
        created_at=result.created_at,
        ip_address=result.ip_address,
        security_risk_level=result.security_risk_level,
        security_categories=result.security_categories,
        compliance_risk_level=result.compliance_risk_level,
        compliance_categories=result.compliance_categories,
        data_risk_level=result.data_risk_level if hasattr(result, 'data_risk_level') else "no_risk",
        data_categories=result.data_categories if hasattr(result, 'data_categories') else [],
        has_image=result.has_image if hasattr(result, 'has_image') else False,
        image_count=result.image_count if hasattr(result, 'image_count') else 0,
        image_paths=result.image_paths if hasattr(result, 'image_paths') else [],
        image_urls=image_urls,
        is_direct_model_access=result.is_direct_model_access if hasattr(result, 'is_direct_model_access') else False,
        source=result.source if hasattr(result, 'source') else None,
        unsafe_segments=result.unsafe_segments if hasattr(result, 'unsafe_segments') and result.unsafe_segments else [],
        doublecheck_result=result.doublecheck_result if hasattr(result, 'doublecheck_result') else None,
        doublecheck_categories=result.doublecheck_categories if hasattr(result, 'doublecheck_categories') else None,
        doublecheck_reasoning=result.doublecheck_reasoning if hasattr(result, 'doublecheck_reasoning') else None,
        application_id=app_id_str,
        application_name=app_info.get("application_name"),
        workspace_id=app_info.get("workspace_id"),
        workspace_name=app_info.get("workspace_name"),
    )


def _generate_image_urls(result) -> list:
    """Generate signed image URLs for a detection result"""
    image_urls = []
    if hasattr(result, 'image_paths') and result.image_paths:
        for image_path in result.image_paths:
            try:
                path_parts = Path(image_path).parts
                filename = path_parts[-1]
                extracted_tenant_id = path_parts[-2]
                signed_url = generate_signed_media_url(
                    tenant_id=extracted_tenant_id,
                    filename=filename,
                    expires_in_seconds=86400
                )
                image_urls.append(signed_url)
            except Exception as e:
                logger.error(f"Failed to generate signed URL for {image_path}: {e}")
    return image_urls


def _build_common_filters(
    tenant_id,
    application_id: Optional[str],
    workspace_id: Optional[str],
    risk_level: Optional[str],
    security_risk_level: Optional[str],
    compliance_risk_level: Optional[str],
    data_risk_level: Optional[str],
    category: Optional[str],
    data_entity_type: Optional[str],
    start_date: Optional[str],
    end_date: Optional[str],
    content_search: Optional[str],
    request_id_search: Optional[str],
    db: Session,
    application_name_search: Optional[str] = None,
    tz_offset: Optional[int] = None,
) -> list:
    """Build common query filters for detection results

    Args:
        tz_offset: Client timezone offset in minutes (from JS getTimezoneOffset()).
                   For UTC+8, this is -480. Used to convert local dates to UTC for filtering.
    """
    filters = []

    # Tenant-level filter: show ALL results for this tenant
    filters.append(DetectionResult.tenant_id == str(tenant_id))

    # Optional application filter
    if application_id:
        filters.append(DetectionResult.application_id == application_id)

    # Optional application name search: fuzzy match on application name
    if application_name_search:
        matching_app_ids = db.query(Application.id).filter(
            Application.tenant_id == str(tenant_id),
            Application.name.ilike(f'%{application_name_search}%')
        ).all()
        matching_ids = [str(a[0]) for a in matching_app_ids]
        if matching_ids:
            filters.append(DetectionResult.application_id.in_(matching_ids))
        else:
            # No matching apps, return empty
            filters.append(DetectionResult.id == -1)

    # Optional workspace filter: find all apps in this workspace, filter by those app IDs
    if workspace_id:
        app_ids = db.query(Application.id).filter(
            Application.workspace_id == workspace_id,
            Application.tenant_id == str(tenant_id),
        ).all()
        app_id_list = [str(a[0]) for a in app_ids]
        if app_id_list:
            filters.append(DetectionResult.application_id.in_(app_id_list))
        else:
            # No apps in workspace, return empty
            filters.append(DetectionResult.id == -1)

    # Risk level filters
    if risk_level:
        if risk_level == "no_risk":
            filters.append(and_(
                DetectionResult.security_risk_level == "no_risk",
                DetectionResult.compliance_risk_level == "no_risk",
                DetectionResult.data_risk_level == "no_risk"
            ))
        elif risk_level == "any_risk":
            filters.append(or_(
                DetectionResult.security_risk_level != "no_risk",
                DetectionResult.compliance_risk_level != "no_risk",
                DetectionResult.data_risk_level != "no_risk"
            ))
        else:
            filters.append(or_(
                DetectionResult.security_risk_level == risk_level,
                DetectionResult.compliance_risk_level == risk_level,
                DetectionResult.data_risk_level == risk_level
            ))

    if security_risk_level:
        if security_risk_level == "any_risk":
            filters.append(DetectionResult.security_risk_level != "no_risk")
        else:
            filters.append(DetectionResult.security_risk_level == security_risk_level)

    if compliance_risk_level:
        if compliance_risk_level == "any_risk":
            filters.append(DetectionResult.compliance_risk_level != "no_risk")
        else:
            filters.append(DetectionResult.compliance_risk_level == compliance_risk_level)

    if data_risk_level:
        if data_risk_level == "any_risk":
            filters.append(DetectionResult.data_risk_level != "no_risk")
        else:
            filters.append(DetectionResult.data_risk_level == data_risk_level)

    if category:
        filters.append(or_(
            cast(DetectionResult.security_categories, JSONB).contains([category]),
            cast(DetectionResult.compliance_categories, JSONB).contains([category]),
            cast(DetectionResult.data_categories, JSONB).contains([category])
        ))

    if data_entity_type:
        filters.append(cast(DetectionResult.data_categories, JSONB).contains([data_entity_type]))

    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        if tz_offset is not None:
            # Convert local midnight to UTC: local_time + tz_offset = UTC
            start_dt = start_dt + timedelta(minutes=tz_offset)
        filters.append(DetectionResult.created_at >= start_dt)

    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        if tz_offset is not None:
            end_dt = end_dt + timedelta(minutes=tz_offset)
        filters.append(DetectionResult.created_at <= end_dt)

    if content_search:
        filters.append(DetectionResult.content.like(f'%{content_search}%'))

    if request_id_search:
        filters.append(DetectionResult.request_id.like(f'%{request_id_search}%'))

    return filters


@router.get("/results")
async def get_detection_results(
    request: Request,
    db: Session = Depends(get_admin_db),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    application_id: Optional[str] = Query(None, description="Filter by application ID"),
    workspace_id: Optional[str] = Query(None, description="Filter by workspace ID"),
    risk_level: Optional[str] = Query(None),
    security_risk_level: Optional[str] = Query(None),
    compliance_risk_level: Optional[str] = Query(None),
    data_risk_level: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    data_entity_type: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    content_search: Optional[str] = Query(None),
    request_id_search: Optional[str] = Query(None),
    application_name_search: Optional[str] = Query(None, description="Fuzzy search by application name"),
    tz_offset: Optional[int] = Query(None, description="Client timezone offset in minutes (JS getTimezoneOffset)"),
):
    """Get detection results (global view - all applications for current tenant)"""
    try:
        current_user = get_current_tenant(request, db)

        filters = _build_common_filters(
            tenant_id=current_user.id,
            application_id=application_id,
            workspace_id=workspace_id,
            risk_level=risk_level,
            security_risk_level=security_risk_level,
            compliance_risk_level=compliance_risk_level,
            data_risk_level=data_risk_level,
            category=category,
            data_entity_type=data_entity_type,
            start_date=start_date,
            end_date=end_date,
            content_search=content_search,
            request_id_search=request_id_search,
            db=db,
            application_name_search=application_name_search,
            tz_offset=tz_offset,
        )

        base_query = db.query(DetectionResult).filter(and_(*filters))
        total = base_query.count()

        offset = (page - 1) * per_page
        results = base_query.order_by(
            DetectionResult.created_at.desc()
        ).offset(offset).limit(per_page).all()

        # Build app/workspace lookup cache
        app_map = _build_app_workspace_cache(db, current_user.id)

        items = []
        for result in results:
            image_urls = _generate_image_urls(result)
            items.append(_enrich_result(result, app_map, image_urls, truncate_content=True))

        pages = (total + per_page - 1) // per_page

        return PaginatedResponse(
            items=items,
            total=total,
            page=page,
            per_page=per_page,
            pages=pages
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get detection results error: {e}")
        raise HTTPException(status_code=500, detail="Failed to get detection results")

@router.get("/results/export")
async def export_detection_results(
    request: Request,
    db: Session = Depends(get_admin_db),
    application_id: Optional[str] = Query(None, description="Filter by application ID"),
    workspace_id: Optional[str] = Query(None, description="Filter by workspace ID"),
    risk_level: Optional[str] = Query(None),
    security_risk_level: Optional[str] = Query(None),
    compliance_risk_level: Optional[str] = Query(None),
    data_risk_level: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    data_entity_type: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    content_search: Optional[str] = Query(None),
    request_id_search: Optional[str] = Query(None),
    application_name_search: Optional[str] = Query(None, description="Fuzzy search by application name"),
    tz_offset: Optional[int] = Query(None, description="Client timezone offset in minutes (JS getTimezoneOffset)"),
):
    """Export detection results to Excel"""
    try:
        current_user = get_current_tenant(request, db)

        filters = _build_common_filters(
            tenant_id=current_user.id,
            application_id=application_id,
            workspace_id=workspace_id,
            risk_level=risk_level,
            security_risk_level=security_risk_level,
            compliance_risk_level=compliance_risk_level,
            data_risk_level=data_risk_level,
            category=category,
            data_entity_type=data_entity_type,
            start_date=start_date,
            end_date=end_date,
            content_search=content_search,
            request_id_search=request_id_search,
            db=db,
            application_name_search=application_name_search,
            tz_offset=tz_offset,
        )

        base_query = db.query(DetectionResult).filter(and_(*filters))
        results = base_query.order_by(
            DetectionResult.created_at.desc()
        ).limit(10000).all()

        # Build app/workspace lookup cache
        app_map = _build_app_workspace_cache(db, current_user.id)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Detection Results"

        headers = [
            "Request ID",
            "Application",
            "Workspace",
            "Detection Content",
            "Prompt Attack Risk",
            "Prompt Attack Categories",
            "Content Compliance Risk",
            "Content Compliance Categories",
            "Data Leak Risk",
            "Data Leak Categories",
            "Suggested Action",
            "Suggested Answer",
            "Hit Keywords",
            "Has Image",
            "Image Count",
            "IP Address",
            "Detection Time"
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_num, result in enumerate(results, 2):
            app_id_str = str(result.application_id) if result.application_id else None
            app_info = app_map.get(app_id_str, {}) if app_id_str else {}

            ws.cell(row=row_num, column=1, value=result.request_id)
            ws.cell(row=row_num, column=2, value=app_info.get("application_name", ""))
            ws.cell(row=row_num, column=3, value=app_info.get("workspace_name", ""))
            ws.cell(row=row_num, column=4, value=result.content)
            ws.cell(row=row_num, column=5, value=result.security_risk_level or "no_risk")
            ws.cell(row=row_num, column=6, value=", ".join(result.security_categories or []))
            ws.cell(row=row_num, column=7, value=result.compliance_risk_level or "no_risk")
            ws.cell(row=row_num, column=8, value=", ".join(result.compliance_categories or []))
            ws.cell(row=row_num, column=9, value=result.data_risk_level or "no_risk")
            ws.cell(row=row_num, column=10, value=", ".join(result.data_categories or []))
            ws.cell(row=row_num, column=11, value=result.suggest_action)
            ws.cell(row=row_num, column=12, value=result.suggest_answer or "")
            ws.cell(row=row_num, column=13, value=", ".join(result.hit_keywords or []))
            ws.cell(row=row_num, column=14, value="Yes" if (hasattr(result, 'has_image') and result.has_image) else "No")
            ws.cell(row=row_num, column=15, value=result.image_count if hasattr(result, 'image_count') else 0)
            ws.cell(row=row_num, column=16, value=result.ip_address or "")
            ws.cell(row=row_num, column=17, value=result.created_at.strftime('%Y-%m-%d %H:%M:%S') if result.created_at else "")

        column_widths = [30, 20, 20, 50, 20, 30, 20, 30, 20, 30, 15, 50, 30, 12, 12, 15, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"detection_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export detection results error: {e}")
        raise HTTPException(status_code=500, detail="Failed to export detection results")

@router.get("/results/{result_id}", response_model=DetectionResultResponse)
async def get_detection_result(result_id: int, request: Request, db: Session = Depends(get_admin_db)):
    """Get single detection result detail (tenant-scoped)"""
    try:
        current_user = get_current_tenant(request, db)

        result = db.query(DetectionResult).filter_by(id=result_id).first()
        if not result:
            raise HTTPException(status_code=404, detail="Detection result not found")

        # Permission check: must belong to current tenant
        if str(result.tenant_id) != str(current_user.id):
            raise HTTPException(status_code=403, detail="Forbidden")

        app_map = _build_app_workspace_cache(db, current_user.id)
        image_urls = _generate_image_urls(result)

        return _enrich_result(result, app_map, image_urls, truncate_content=False)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get detection result error: {e}")
        raise HTTPException(status_code=500, detail="Failed to get detection result")


@router.post("/results/{result_id}/extract-segments")
async def extract_unsafe_segments(result_id: int, request: Request, db: Session = Depends(get_admin_db)):
    """
    On-demand extraction of unsafe content segments.
    Called when user opens log detail for a result with compliance/security risk.
    Only runs for results with compliance or security risk (not data-only or safe).
    Results are cached in DB after first extraction.
    """
    try:
        current_user = get_current_tenant(request, db)

        result = db.query(DetectionResult).filter_by(id=result_id).first()
        if not result:
            raise HTTPException(status_code=404, detail="Detection result not found")

        # Permission check
        if str(result.tenant_id) != str(current_user.id):
            raise HTTPException(status_code=403, detail="Forbidden")

        # Return cached result if already extracted
        if result.unsafe_segments and len(result.unsafe_segments) > 0:
            return {"unsafe_segments": result.unsafe_segments}

        # Only extract for compliance/security risks (not data-only or safe)
        has_compliance_risk = result.compliance_risk_level and result.compliance_risk_level != "no_risk"
        has_security_risk = result.security_risk_level and result.security_risk_level != "no_risk"

        if not has_compliance_risk and not has_security_risk:
            return {"unsafe_segments": []}

        # Build matched categories from the stored result
        matched_categories = []
        # Extract scanner tags from compliance/security categories
        # Try matched_scanner_tags from model_response or reconstruct from categories
        if result.security_categories:
            matched_categories.append("S9")  # Security = Prompt Attacks
        if result.compliance_categories:
            # Reverse-lookup category names to tags
            from services.guardrail_service import CATEGORY_NAMES
            name_to_tag = {v: k for k, v in CATEGORY_NAMES.items()}
            for cat_name in result.compliance_categories:
                tag = name_to_tag.get(cat_name)
                if tag:
                    matched_categories.append(tag)

        if not matched_categories:
            return {"unsafe_segments": []}

        # Get scanner definitions for context if application_id is available
        scanner_defs = None
        if result.application_id:
            try:
                from services.scanner_config_service import ScannerConfigService
                from services.scanner_detection_service import ScannerDetectionService
                scs = ScannerConfigService(db)
                all_scanners = scs.get_application_scanners(
                    application_id=result.application_id,
                    tenant_id=result.tenant_id,
                    include_disabled=False
                )
                genai_scanners = [s for s in all_scanners if s['scanner_type'] == 'genai']
                if genai_scanners:
                    temp_sds = ScannerDetectionService(db)
                    scanner_defs = temp_sds._prepare_scanner_definitions(genai_scanners)
            except Exception as e:
                logger.warning(f"Failed to get scanner definitions: {e}")

        # Run extraction
        from services.model_service import model_service
        unsafe_segments = await model_service.extract_unsafe_segments(
            content=result.content,
            matched_categories=matched_categories,
            scanner_definitions=scanner_defs
        )

        # Save to DB for caching
        result.unsafe_segments = unsafe_segments
        db.commit()

        return {"unsafe_segments": unsafe_segments}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Extract unsafe segments error: {e}")
        raise HTTPException(status_code=500, detail="Failed to extract unsafe segments")
